# Author: Adam Vengroff
# Description: This software processes serial data in order to evaluate
#              welding gun operation, graph data in real-time as well as
#              as stream/record video

# RPi peripherals imports
import picamera

# Networking imports
import socket

# Excel imports
import openpyxl

# Audio imports
import os

# Custom-written module imports
from DualOutput import DualOutput
from SheetsThread import SheetsThread

# Google Project Api Imports
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Multi-threading imports
import threading

# Read USB data
import serial
import time

# Time stamping
import datetime

# Data simulation
import math

# GUI
from tkinter import *

# Graphing
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.animation as animation

# For speed improvements
from collections import deque

# Define x limit
xLim = 30

# Create client to interact with Google Drive Api
scope = ['https://spreadsheets.google.com/feeds']
creds = ServiceAccountCredentials.from_json_keyfile_name('/home/pi/Downloads/client.json', scope)
client = gspread.authorize(creds)

# Find workbook and open first sheet
sheet = client.open("WeldingReadIn").sheet1

# Open up welding parameter data worksheet
wb = openpyxl.load_workbook('/home/pi/Downloads/WeldingParameters.xlsx')
ws = wb["MIG"]

# Constants definite excel sheet size
CONST_START_ROW = 5
CONST_END_ROW = 48
CONST_START_COLUMN = 4
CONST_END_COLUMN = 15

# Constants
UPDATE_FREQ = 150 # How many iterations before writing to Google Sheets
BAUD_RATE = 57600 # Communication speed with Arduino
SESSION_LENGTH = 3600 # How many seconds to stream video

# Create 2D Array to hold parameter data
w = CONST_END_ROW - CONST_START_ROW + 1
h = CONST_END_ROW - CONST_START_ROW + 1
weldingParameters = [['-1' for x in range(w)] for y in range(h)] # weldingParameters[r][c]

# Welding Parameters
distance = [0]
current = [0]
angle = [0]
accFB = [0]
accLR = [0]
timestamp = []
parameterList = []
parameterList.extend((angle, distance, accFB, accLR, current, timestamp))

# Parameter defines for Table Look-up
metal = ['0', '1']
transfer = ['0', '1']
thickness = ['00', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10']

# Welding settings
metalIndex = int(sheet.cell(3, 3).value)
transferIndex = int(sheet.cell(4, 3).value)
thicknessIndex = int(sheet.cell(5, 3).value)
displayMode = int(sheet.cell(6, 3).value)
audioIndicator = int(sheet.cell(7, 3).value)

# USB set-up
ser = serial.Serial('/dev/ttyUSB0', BAUD_RATE)

# Load data from excel sheet
for r in range(CONST_START_ROW, CONST_END_ROW + 1):
    for angPlot in range(CONST_START_COLUMN, CONST_END_COLUMN + 1):
        cellStr = ws.cell(row = r, column = angPlot)
        weldingParameters[r - CONST_START_ROW][angPlot - CONST_START_COLUMN] = cellStr.value

# Extract ID Column
idCol = [item[11] for item in weldingParameters]

# Create ID
id = metal[metalIndex] + transfer[transferIndex] + thickness[thicknessIndex]

# Get table index from id
tableIndex = idCol.index(id)

# Load welding parameter information from table
currentLow = weldingParameters[tableIndex][3]
currentHigh = weldingParameters[tableIndex][4]
shieldingGas = weldingParameters[tableIndex][5]
voltageRange = weldingParameters[tableIndex][6]
thinWireSize = weldingParameters[tableIndex][7]
thinWFS = weldingParameters[tableIndex][8]
thickWireSize = weldingParameters[tableIndex][9]
thickWFS = weldingParameters[tableIndex][10]

# Angle ranges for different welding types
angleLow = [75, 40, 60]
angleHigh = [105, 50, 70]
angleType = ['Butt Weld', 'T-Joint', 'Lap Joint']

# Data plotting variables
curDataGood = deque([None] * xLim, maxlen = xLim)
distDataGood = deque([None] * xLim, maxlen = xLim)
angDataGood = deque([None] * xLim, maxlen = xLim)
accDataGood = deque([None] * xLim, maxlen = xLim)

curDataBad = deque([None] * xLim, maxlen = xLim)
distDataBad = deque([None] * xLim, maxlen = xLim)
angDataBad = deque([None] * xLim, maxlen = xLim)
accDataBad = deque([None] * xLim, maxlen = xLim)

# Parameter range set-up
curRange = [198, 202]
distRange = [10, 12]
angRange = [85, 95]
accRange = [-0.2, 0.2]

# GUI Set-Up
outStr = ""
ROOT = Tk()
f = Frame(ROOT)
ROOT.attributes("-fullscreen", True)
displayMode = 2
var = StringVar()
var.set("test")
LABEL = Label(f, textvariable = var)

# Graph animation
def animate(i):
    lineCurGood.set_data(range(len(curDataGood)), curDataGood)
    lineCurBad.set_data(range(len(curDataBad)), curDataBad)
    lineDistGood.set_data(range(len(distDataGood)), distDataGood)
    lineDistBad.set_data(range(len(distDataBad)), distDataBad)
    lineAngGood.set_data(range(len(angDataGood)), angDataGood)
    lineAngBad.set_data(range(len(angDataBad)), angDataBad)
    lineAccGood.set_data(range(len(accDataGood)), accDataGood)
    lineAccBad.set_data(range(len(accDataBad)), accDataBad)
    return lineCurGood, lineCurBad, lineDistGood, lineDistBad,\
           lineAngGood, lineAngBad, lineAccGood, lineAccBad

# For using blit to improve efficiency by only re-drawing
# these items during the animation
def animateinit():
    return lineCurGood, lineCurBad, lineDistGood, lineDistBad,\
           lineAngGood, lineAngBad, lineAccGood, lineAccBad

# Adds data to end of appropriate data array using parameter type
# and determining of data point is within target range, removes old data
# from start of the array to maintain constant array size
def putData(data, param):
    if param == 0:
        curDataGood.append(data)
        curDataGood.popleft()
        if curRange[0] < data < curRange[1]:
            curDataBad.append(None)
            curDataBad.popleft()
        else:
            curDataBad.append(data)
            curDataBad.popleft()

    elif param == 1:
        distDataGood.append(data)
        distDataGood.popleft()
        if distRange[0] < data < distRange[1]:
            distDataBad.append(None)
            distDataBad.popleft()
        else:
            distDataBad.append(data)
            distDataBad.popleft()

    elif param == 2:
        angDataGood.append(data)
        angDataGood.popleft()
        if angRange[0] < data < angRange[1]:
            angDataBad.append(None)
            angDataBad.popleft()
        else:
            angDataBad.append(data)
            angDataBad.popleft()

    elif param == 3:
        accDataGood.append(data)
        accDataGood.popleft()
        if accRange[0] < data < accRange[1]:
            accDataBad.append(None)
            accDataBad.popleft()
        else:
            accDataBad.append(data)
            accDataBad.popleft()


# Push update to frame
def updateFrame():
    LABEL.config(font=("Courier", 55))
    LABEL.pack()
    f.pack()
    ROOT.update()

# RunThread is responsible for video recording/streaming as well as recieving
# data from Arduino and analyzing it
def RunThread():
    with picamera.PiCamera() as camera:
        camera.resolution = (1920, 1080)
        camera.framerate = 30
        camera.brightness = 55
        camera.annotate_text_size = 16
        i = 0
        param = 0

        streamFlag = int(sheet.cell(6, 3).value)

        try:
            tsString = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')

            if ((streamFlag == 1) or (streamFlag == 3)):
                server_socket = socket.socket()

                try:
                    server_socket.bind(('0.0.0.0', 8000))
                    print('Waiting on connection via port 8000')
                except:
                    server_socket.bind(('0.0.0.0', 8001))
                    print('Waiting on connection via port 8001')

                server_socket.listen(0)
                connection = server_socket.accept()[0].makefile('wb')

                if (streamFlag == 1):
                    print('Stream Starting')
                    camera.start_recording(connection, format='h264')
                else:
                    print('Stream and Recording Starting')
                    customOutput = DualOutput('/home/pi/Downloads/' + tsString + '.h264', connection)
                    camera.start_recording(customOutput, format='h264')

            elif (streamFlag == 2):
                print('Recording Starting')
                camera.start_recording('/home/pi/Downloads/' + tsString + '.h264')

            timeout = time.time() + SESSION_LENGTH

            # track number of iterations
            index = 0
            indexOffset = 0

            while time.time() < timeout:
                data = str(ser.readline())

                try:
                    rawData, measurement = data.split(":")
                    measurementValue = measurement.split("\\", 1)[0]
                except:
                    rawData = ""
                    print('Serial Read Error')

                if "C" in rawData:
                    param = 0
                    measurementValue = 200 + (5 * math.sin(index))
                    current.append(round(float(measurementValue), 3))

                elif "D" in rawData:
                    param = 1
                    distance.append(measurementValue)

                elif "angLR" in rawData:
                    angle.append(measurementValue)
                    param = 2
                    if float(angle[-1]) > 75:
                        i = 0
                    elif float(angle[-1]) < 55:
                        i = 2
                    else:
                        i = 1

                elif "accLR" in rawData:
                    accLR.append(round(float(measurementValue), 3))
                    param = 3

                elif "accFB" in rawData:
                    accFB.append(round(float(measurementValue), 3))
                    param = 4

                elif "angFB" in rawData:
                    param = 4

                parsedData = measurementValue

                index = index + 1
                timestamp.append(datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S'))

                currentStr = 'Current: ' + str(currentLow) + ' < ' + str(current[-1])[:3] + ' < ' + str(currentHigh) + '\n'
                distanceStr = 'Distance: ' + str(distRange[0]) + ' < ' + str(distance[-1])[:5] + ' < ' + str(distRange[1]) + '\n'
                angleStr = 'Angle (' + str(angleType[i]) + '): ' + str(angleLow[i]) + ' < ' + str(angle[-1])[:3] + ' < ' + str(angleHigh[i]) + '\n'
                accStr = 'LR Acc: ' + str(accLR[-1])[:3] + ' FB Acc: ' + str(accFB[-1])[:3]
                outStr = currentStr + distanceStr + angleStr + accStr
                camera.annotate_text = outStr

                if displayMode == 1:
                    var.set(outStr)
                    updateFrame()
                elif displayMode == 2:
                    putData(float(parsedData), param)

                # Call thread to write to Google sheets and update GUI every UPDATE_FREQ samples
                if (index % UPDATE_FREQ == 0):
                    newSheetThread = threading.Thread(target=SheetsThread,
                                                      args=(indexOffset, UPDATE_FREQ, parameterList, sheet,))
                    newSheetThread.start()
                    indexOffset = index

            camera.stop_recording()

        finally:
            if (streamFlag == 1 or streamFlag == 3):
                connection.close()
                server_socket.close()
                print('Stream Ended')
            else:
                print('Recording Complete')

        f.destroy()
        ROOT.destroy()

# Real-time graph set-up
fig = Figure(figsize = (12, 7), dpi = 100)

curPlot = fig.add_subplot(221)
distPlot = fig.add_subplot(222)
angPlot = fig.add_subplot(223)
accPlot = fig.add_subplot(224)

curPlot.set_xlim([0, xLim])
distPlot.set_xlim([0, xLim])
angPlot.set_xlim([0, xLim])
accPlot.set_xlim([0, xLim])

curPlot.set_ylim([-220, 220])
distPlot.set_ylim([0, 15])
angPlot.set_ylim([-180, 180])
accPlot.set_ylim([-1, 1])

lineCurGood, = curPlot.plot([0], [0], color = 'black')
lineCurBad, = curPlot.plot([0], [0],color = 'red')
lineDistGood, = distPlot.plot([0], [0],color = 'black')
lineDistBad, = distPlot.plot([0], [0],color = 'red')
lineAngGood, = angPlot.plot([0], [0],color = 'black')
lineAngBad, = angPlot.plot([0], [0],color = 'red')
lineAccGood, = accPlot.plot([0], [0],color = 'black')
lineAccBad, = accPlot.plot([0], [0],color = 'red')

curPlot.axis('off')
curPlot.set_title("Current", fontsize = 25)
distPlot.axis('off')
distPlot.set_title("Distance", fontsize = 25)
angPlot.axis('off')
angPlot.set_title("Angle", fontsize = 25)
accPlot.axis('off')
accPlot.set_title("Acceleration", fontsize = 25)

# GUI set-up
canvas = FigureCanvasTkAgg(fig, master=ROOT)
canvas.get_tk_widget().pack(side=BOTTOM, fill=BOTH, expand=True)
canvas._tkcanvas.pack(side=TOP, fill=BOTH, expand=True)

Frame.pack(f)

ani = animation.FuncAnimation(fig, animate, interval=40, init_func = animateinit, blit = True)
ROOT.config(cursor="none")
print('Activating GUI')

newRunThread = threading.Thread(target=RunThread, args=())
newRunThread.start()

ROOT.mainloop()



